"""Excel report generator.

Produces a multi-sheet .xlsx report file containing:
  - Summary overview sheet
  - Per-column analysis sheet
  - Data preview sheet with type-annotated headers
  - Data quality sheet
"""

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from task_manager.excel_report.analyzer import AnalysisResult, ColumnStats
from task_manager.excel_report.importer import DataType, ImportResult


# Color fills for data types (without '#' prefix for openpyxl)
TYPE_FILLS = {
    DataType.INTEGER.value: "2196F3",
    DataType.FLOAT.value: "03A9F4",
    DataType.PERCENTAGE.value: "00BCD4",
    DataType.CURRENCY.value: "4CAF50",
    DataType.TEXT.value: "FF9800",
    DataType.BOOLEAN.value: "9C27B0",
    DataType.DATE.value: "E91E63",
    DataType.TIME.value: "F44336",
    DataType.DATETIME.value: "E91E63",
    DataType.EMAIL.value: "795548",
    DataType.URL.value: "607D8B",
    DataType.PHONE.value: "009688",
    DataType.EMPTY.value: "9E9E9E",
    DataType.MIXED.value: "FF5722",
}


def _header_font() -> Font:
    return Font(bold=True, color="FFFFFF", size=11)


def _header_fill() -> PatternFill:
    return PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")


def _type_fill(type_name: str) -> PatternFill:
    color = TYPE_FILLS.get(type_name, "999999")
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _thin_border() -> Border:
    side = Side(style="thin", color="DADCE0")
    return Border(left=side, right=side, top=side, bottom=side)


def _style_header_row(ws, row: int, col_count: int) -> None:
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()


class ExcelReportGenerator:
    """Generates a multi-sheet Excel report with styling."""

    def __init__(self, preview_rows: int = 50):
        self.preview_rows = preview_rows

    def generate(
        self,
        import_result: ImportResult,
        analysis_result: AnalysisResult,
        output_path: str,
    ) -> str:
        """Generate an Excel report and save to output_path.

        Args:
            import_result: Data from the importer.
            analysis_result: Statistics from the analyzer.
            output_path: File path for the .xlsx output.

        Returns:
            The output file path.
        """
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)

        self._create_summary_sheet(wb, import_result, analysis_result)

        for sheet_data, sheet_analysis in zip(
            import_result.sheets, analysis_result.sheets
        ):
            self._create_column_analysis_sheet(wb, sheet_analysis)
            self._create_data_preview_sheet(wb, sheet_data, sheet_analysis)
            self._create_quality_sheet(wb, sheet_analysis)

        wb.save(output_path)
        return output_path

    def _create_summary_sheet(
        self,
        wb: Workbook,
        import_result: ImportResult,
        analysis_result: AnalysisResult,
    ) -> None:
        """Create the overview summary sheet."""
        ws = wb.create_sheet("Summary")

        # Title
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"Excel Import Report: {import_result.file_name}"
        title_cell.font = Font(bold=True, size=16, color="1A73E8")
        title_cell.alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:F2")
        ws["A2"].value = f"Generated: {analysis_result.analysis_time}"
        ws["A2"].font = Font(color="5F6368", size=10)
        ws["A2"].alignment = Alignment(horizontal="center")

        # Overall stats
        row = 4
        stats_headers = [
            "Metric",
            "Value",
        ]
        for col, header in enumerate(stats_headers, 1):
            ws.cell(row=row, column=col, value=header)
        _style_header_row(ws, row, len(stats_headers))

        stats = [
            ("Total Sheets", len(import_result.sheets)),
            ("Total Data Rows", import_result.total_rows),
            ("Total Columns", import_result.total_columns),
            (
                "Overall Data Quality Score",
                f"{round(analysis_result.overall_quality_score, 1)}%",
            ),
        ]

        for i, (label, value) in enumerate(stats, 1):
            ws.cell(row=row + i, column=1, value=label).border = _thin_border()
            cell = ws.cell(row=row + i, column=2, value=value)
            cell.border = _thin_border()
            cell.font = Font(bold=True)

        # Per-sheet summary
        row = row + len(stats) + 3
        ws.cell(row=row, column=1, value="Sheet Summary")
        ws.cell(row=row, column=1).font = Font(bold=True, size=13, color="1A73E8")

        row += 1
        sheet_headers = [
            "Sheet Name",
            "Rows",
            "Columns",
            "Quality Score",
            "Duplicate Rows",
            "Empty Rows",
        ]
        for col, header in enumerate(sheet_headers, 1):
            ws.cell(row=row, column=col, value=header)
        _style_header_row(ws, row, len(sheet_headers))

        for i, sa in enumerate(analysis_result.sheets, 1):
            r = row + i
            values = [
                sa.sheet_name,
                sa.row_count,
                sa.col_count,
                f"{round(sa.data_quality_score, 1)}%",
                sa.duplicate_rows,
                sa.completely_empty_rows,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = _thin_border()

        # Type distribution across all sheets
        row = row + len(analysis_result.sheets) + 3
        ws.cell(row=row, column=1, value="Data Type Distribution")
        ws.cell(row=row, column=1).font = Font(bold=True, size=13, color="1A73E8")

        row += 1
        type_headers = ["Data Type", "Column Count"]
        for col, header in enumerate(type_headers, 1):
            ws.cell(row=row, column=col, value=header)
        _style_header_row(ws, row, len(type_headers))

        # Aggregate type counts
        type_totals: dict[str, int] = {}
        for sa in analysis_result.sheets:
            for type_name, count in sa.type_summary.items():
                type_totals[type_name] = type_totals.get(type_name, 0) + count

        for i, (type_name, count) in enumerate(
            sorted(type_totals.items(), key=lambda x: -x[1]), 1
        ):
            r = row + i
            type_cell = ws.cell(row=r, column=1, value=type_name)
            type_cell.fill = _type_fill(type_name)
            type_cell.font = Font(color="FFFFFF", bold=True)
            type_cell.border = _thin_border()
            count_cell = ws.cell(row=r, column=2, value=count)
            count_cell.border = _thin_border()
            count_cell.font = Font(bold=True)

        # Auto-width
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 22

    def _create_column_analysis_sheet(
        self, wb: Workbook, sheet_analysis: "AnalysisResult"
    ) -> None:
        """Create per-column analysis sheet."""
        name = f"Analysis - {sheet_analysis.sheet_name}"[:31]
        ws = wb.create_sheet(name)

        headers = [
            "Column",
            "Data Type",
            "Total",
            "Non-Empty",
            "Empty",
            "Unique",
            "Completeness",
            "Duplicates",
            "Min",
            "Max",
            "Mean",
            "Median",
            "Std Dev",
            "Sum",
            "Outliers",
            "Top Value",
            "Top Count",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        _style_header_row(ws, 1, len(headers))

        for i, cs in enumerate(sheet_analysis.column_stats, 1):
            r = i + 1
            ws.cell(row=r, column=1, value=cs.header).border = _thin_border()

            type_cell = ws.cell(row=r, column=2, value=cs.data_type.value)
            type_cell.fill = _type_fill(cs.data_type.value)
            type_cell.font = Font(color="FFFFFF", bold=True)
            type_cell.border = _thin_border()

            ws.cell(row=r, column=3, value=cs.total_count).border = _thin_border()
            ws.cell(row=r, column=4, value=cs.non_empty_count).border = _thin_border()
            ws.cell(row=r, column=5, value=cs.empty_count).border = _thin_border()
            ws.cell(row=r, column=6, value=cs.unique_count).border = _thin_border()

            comp_cell = ws.cell(
                row=r, column=7, value=f"{round(cs.completeness, 1)}%"
            )
            comp_cell.border = _thin_border()
            if cs.completeness >= 90:
                comp_cell.font = Font(color="34A853", bold=True)
            elif cs.completeness >= 70:
                comp_cell.font = Font(color="FBBC04", bold=True)
            else:
                comp_cell.font = Font(color="EA4335", bold=True)

            ws.cell(row=r, column=8, value=cs.duplicate_count).border = _thin_border()

            # Numeric stats
            if cs.min_val is not None:
                ws.cell(row=r, column=9, value=cs.min_val).border = _thin_border()
            else:
                ws.cell(row=r, column=9, value="-").border = _thin_border()
            if cs.max_val is not None:
                ws.cell(row=r, column=10, value=cs.max_val).border = _thin_border()
            else:
                ws.cell(row=r, column=10, value="-").border = _thin_border()
            if cs.mean_val is not None:
                ws.cell(
                    row=r, column=11, value=round(cs.mean_val, 4)
                ).border = _thin_border()
            else:
                ws.cell(row=r, column=11, value="-").border = _thin_border()
            if cs.median_val is not None:
                ws.cell(row=r, column=12, value=cs.median_val).border = _thin_border()
            else:
                ws.cell(row=r, column=12, value="-").border = _thin_border()
            if cs.std_dev is not None:
                ws.cell(
                    row=r, column=13, value=round(cs.std_dev, 4)
                ).border = _thin_border()
            else:
                ws.cell(row=r, column=13, value="-").border = _thin_border()
            if cs.sum_val is not None:
                ws.cell(
                    row=r, column=14, value=round(cs.sum_val, 4)
                ).border = _thin_border()
            else:
                ws.cell(row=r, column=14, value="-").border = _thin_border()

            if cs.has_outliers:
                outlier_cell = ws.cell(row=r, column=15, value=cs.outlier_count)
                outlier_cell.font = Font(color="EA4335", bold=True)
                outlier_cell.border = _thin_border()
            else:
                ws.cell(row=r, column=15, value=0).border = _thin_border()

            # Top value
            if cs.top_values:
                ws.cell(
                    row=r, column=16, value=str(cs.top_values[0][0])
                ).border = _thin_border()
                ws.cell(
                    row=r, column=17, value=cs.top_values[0][1]
                ).border = _thin_border()
            else:
                ws.cell(row=r, column=16, value="-").border = _thin_border()
                ws.cell(row=r, column=17, value="-").border = _thin_border()

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

    def _create_data_preview_sheet(
        self, wb: Workbook, sheet_data: Any, sheet_analysis: Any
    ) -> None:
        """Create a data preview sheet with type-annotated headers."""
        name = f"Data - {sheet_data.name}"[:31]
        ws = wb.create_sheet(name)

        num_cols = len(sheet_data.headers)

        # Row 1: Headers
        for col, header in enumerate(sheet_data.headers, 1):
            ws.cell(row=1, column=col, value=header)
        _style_header_row(ws, 1, num_cols)

        # Row 2: Data types
        for col, col_info in enumerate(sheet_data.columns, 1):
            cell = ws.cell(row=2, column=col, value=col_info.detected_type.value)
            cell.fill = _type_fill(col_info.detected_type.value)
            cell.font = Font(color="FFFFFF", size=9, bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = _thin_border()

        # Data rows
        for row_idx, row_data in enumerate(
            sheet_data.rows[: self.preview_rows], start=3
        ):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=str(value) if value is not None else "",
                )
                cell.border = _thin_border()

        # Auto-width
        for col in range(1, num_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_quality_sheet(self, wb: Workbook, sheet_analysis: Any) -> None:
        """Create a data quality assessment sheet."""
        name = f"Quality - {sheet_analysis.sheet_name}"[:31]
        ws = wb.create_sheet(name)

        # Title
        ws.merge_cells("A1:E1")
        title = ws["A1"]
        title.value = f"Data Quality Report: {sheet_analysis.sheet_name}"
        title.font = Font(bold=True, size=14, color="1A73E8")

        ws.merge_cells("A2:E2")
        score_cell = ws["A2"]
        score_cell.value = (
            f"Overall Quality Score: {round(sheet_analysis.data_quality_score, 1)}%"
        )
        if sheet_analysis.data_quality_score >= 80:
            score_cell.font = Font(bold=True, size=12, color="34A853")
        elif sheet_analysis.data_quality_score >= 60:
            score_cell.font = Font(bold=True, size=12, color="FBBC04")
        else:
            score_cell.font = Font(bold=True, size=12, color="EA4335")

        # Quality details per column
        row = 4
        headers = [
            "Column",
            "Completeness",
            "Unique Values",
            "Duplicates",
            "Quality Rating",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        _style_header_row(ws, row, len(headers))

        for i, cs in enumerate(sheet_analysis.column_stats, 1):
            r = row + i
            ws.cell(row=r, column=1, value=cs.header).border = _thin_border()

            comp_cell = ws.cell(
                row=r, column=2, value=f"{round(cs.completeness, 1)}%"
            )
            comp_cell.border = _thin_border()

            ws.cell(row=r, column=3, value=cs.unique_count).border = _thin_border()
            ws.cell(row=r, column=4, value=cs.duplicate_count).border = _thin_border()

            # Rating
            if cs.completeness >= 95:
                rating = "Excellent"
                color = "34A853"
            elif cs.completeness >= 80:
                rating = "Good"
                color = "1A73E8"
            elif cs.completeness >= 60:
                rating = "Fair"
                color = "FBBC04"
            else:
                rating = "Poor"
                color = "EA4335"

            rating_cell = ws.cell(row=r, column=5, value=rating)
            rating_cell.font = Font(bold=True, color=color)
            rating_cell.border = _thin_border()

        # Issues section
        row = row + len(sheet_analysis.column_stats) + 3
        ws.cell(row=row, column=1, value="Data Issues Detected")
        ws.cell(row=row, column=1).font = Font(bold=True, size=13, color="1A73E8")

        row += 1
        issue_headers = ["Issue Type", "Details", "Severity"]
        for col, header in enumerate(issue_headers, 1):
            ws.cell(row=row, column=col, value=header)
        _style_header_row(ws, row, len(issue_headers))

        issues: list[tuple[str, str, str]] = []

        if sheet_analysis.duplicate_rows > 0:
            issues.append((
                "Duplicate Rows",
                f"{sheet_analysis.duplicate_rows} duplicate row(s) found",
                "Warning",
            ))

        if sheet_analysis.completely_empty_rows > 0:
            issues.append((
                "Empty Rows",
                f"{sheet_analysis.completely_empty_rows} completely empty row(s)",
                "Info",
            ))

        for cs in sheet_analysis.column_stats:
            if cs.completeness < 50:
                issues.append((
                    "Low Completeness",
                    f"Column '{cs.header}' is only {round(cs.completeness, 1)}% complete",
                    "Critical",
                ))
            elif cs.completeness < 80:
                issues.append((
                    "Moderate Completeness",
                    f"Column '{cs.header}' is {round(cs.completeness, 1)}% complete",
                    "Warning",
                ))

            if cs.has_outliers:
                issues.append((
                    "Outliers Detected",
                    f"Column '{cs.header}' has {cs.outlier_count} outlier(s)",
                    "Info",
                ))

        if not issues:
            issues.append(("None", "No data quality issues detected", "OK"))

        severity_colors = {
            "Critical": "EA4335",
            "Warning": "FBBC04",
            "Info": "1A73E8",
            "OK": "34A853",
        }

        for i, (issue, detail, severity) in enumerate(issues, 1):
            r = row + i
            ws.cell(row=r, column=1, value=issue).border = _thin_border()
            ws.cell(row=r, column=2, value=detail).border = _thin_border()
            sev_cell = ws.cell(row=r, column=3, value=severity)
            sev_cell.font = Font(bold=True, color=severity_colors.get(severity, "000000"))
            sev_cell.border = _thin_border()

        # Auto-width
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 24
