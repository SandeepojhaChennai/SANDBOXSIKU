"""Comprehensive tests for the Excel Import & Report system.

Tests cover:
  - Excel importer with data type detection
  - Data analyzer statistics computation
  - HTML report generation
  - Excel report generation
  - CLI integration
"""

import os
import tempfile
from datetime import datetime, date

import pytest
from openpyxl import Workbook

from task_manager.excel_report.importer import (
    DataType,
    ExcelImporter,
    _detect_cell_type,
    _resolve_column_type,
)
from task_manager.excel_report.analyzer import DataAnalyzer
from task_manager.excel_report.report_html import HTMLReportGenerator
from task_manager.excel_report.report_excel import ExcelReportGenerator


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _create_test_workbook(
    sheets: dict[str, tuple[list[str], list[list]]],
) -> str:
    """Create a temporary .xlsx file with the given sheets.

    Args:
        sheets: Mapping of sheet_name -> (headers, rows).

    Returns:
        Path to the temporary .xlsx file.
    """
    wb = Workbook()
    default = wb.active
    first = True
    for sheet_name, (headers, rows) in sheets.items():
        if first:
            ws = default
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for row in rows:
            ws.append(row)

    path = tempfile.mktemp(suffix=".xlsx")
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def sample_xlsx():
    """Create a sample Excel file with diverse data types."""
    headers = [
        "ID",
        "Name",
        "Email",
        "Age",
        "Salary",
        "Active",
        "Join Date",
        "Score",
        "Website",
        "Phone",
    ]
    rows = [
        [
            1,
            "Alice Smith",
            "alice@example.com",
            30,
            75000.50,
            True,
            datetime(2020, 1, 15),
            95.5,
            "https://alice.dev",
            "+1-555-0101",
        ],
        [
            2,
            "Bob Jones",
            "bob@example.com",
            25,
            62000.00,
            False,
            datetime(2021, 3, 22),
            88.0,
            "https://bob.io",
            "+1-555-0102",
        ],
        [
            3,
            "Charlie Brown",
            "charlie@example.com",
            35,
            90000.75,
            True,
            datetime(2019, 7, 10),
            92.3,
            "https://charlie.net",
            "+1-555-0103",
        ],
        [
            4,
            "Diana Prince",
            "diana@example.com",
            28,
            85000.25,
            True,
            datetime(2022, 11, 5),
            97.1,
            "https://diana.org",
            "+1-555-0104",
        ],
        [
            5,
            "Eve Wilson",
            "eve@example.com",
            40,
            105000.00,
            False,
            datetime(2018, 6, 1),
            91.7,
            "https://eve.com",
            "+1-555-0105",
        ],
    ]
    path = _create_test_workbook({"Employees": (headers, rows)})
    yield path
    os.unlink(path)


@pytest.fixture
def multi_sheet_xlsx():
    """Create a workbook with multiple sheets and mixed data."""
    headers1 = ["Product", "Price", "Quantity", "In Stock"]
    rows1 = [
        ["Widget", 9.99, 100, True],
        ["Gadget", 24.50, 50, True],
        ["Doohickey", 4.75, 200, False],
        ["Thingamajig", 15.00, 0, False],
    ]

    headers2 = ["Date", "Revenue", "Expenses", "Notes"]
    rows2 = [
        [datetime(2024, 1, 1), 50000, 30000, "Good month"],
        [datetime(2024, 2, 1), 45000, 28000, "Slightly down"],
        [datetime(2024, 3, 1), 60000, 35000, "Best quarter"],
    ]

    path = _create_test_workbook({
        "Products": (headers1, rows1),
        "Financials": (headers2, rows2),
    })
    yield path
    os.unlink(path)


@pytest.fixture
def sparse_xlsx():
    """Create a workbook with empty cells and duplicates."""
    headers = ["A", "B", "C"]
    rows = [
        [1, "hello", None],
        [None, "hello", None],
        [3, None, None],
        [1, "hello", None],
        [None, None, None],
    ]
    path = _create_test_workbook({"Sparse": (headers, rows)})
    yield path
    os.unlink(path)


# ---------------------------------------------------------------------------
# Cell type detection tests
# ---------------------------------------------------------------------------


class TestCellTypeDetection:
    def test_none_is_empty(self):
        assert _detect_cell_type(None) == DataType.EMPTY

    def test_empty_string_is_empty(self):
        assert _detect_cell_type("") == DataType.EMPTY
        assert _detect_cell_type("   ") == DataType.EMPTY

    def test_bool_detection(self):
        assert _detect_cell_type(True) == DataType.BOOLEAN
        assert _detect_cell_type(False) == DataType.BOOLEAN

    def test_bool_strings(self):
        assert _detect_cell_type("true") == DataType.BOOLEAN
        assert _detect_cell_type("False") == DataType.BOOLEAN
        assert _detect_cell_type("yes") == DataType.BOOLEAN
        assert _detect_cell_type("NO") == DataType.BOOLEAN

    def test_integer_detection(self):
        assert _detect_cell_type(42) == DataType.INTEGER
        assert _detect_cell_type(0) == DataType.INTEGER
        assert _detect_cell_type(-5) == DataType.INTEGER

    def test_integer_string(self):
        assert _detect_cell_type("42") == DataType.INTEGER
        assert _detect_cell_type("1,000") == DataType.INTEGER

    def test_float_detection(self):
        assert _detect_cell_type(3.14) == DataType.FLOAT
        assert _detect_cell_type(0.0) == DataType.FLOAT

    def test_float_string(self):
        assert _detect_cell_type("3.14") == DataType.FLOAT
        assert _detect_cell_type("1,234.56") == DataType.FLOAT

    def test_datetime_detection(self):
        assert _detect_cell_type(datetime(2024, 1, 1, 12, 30)) == DataType.DATETIME

    def test_date_detection(self):
        assert _detect_cell_type(date(2024, 1, 1)) == DataType.DATE

    def test_email_detection(self):
        assert _detect_cell_type("user@example.com") == DataType.EMAIL
        assert _detect_cell_type("test.name+tag@domain.co") == DataType.EMAIL

    def test_url_detection(self):
        assert _detect_cell_type("https://example.com") == DataType.URL
        assert _detect_cell_type("http://foo.bar/path") == DataType.URL

    def test_phone_detection(self):
        assert _detect_cell_type("+1-555-0101") == DataType.PHONE
        assert _detect_cell_type("(555) 123-4567") == DataType.PHONE

    def test_percentage_detection(self):
        assert _detect_cell_type("85%") == DataType.PERCENTAGE
        assert _detect_cell_type("12.5%") == DataType.PERCENTAGE
        assert _detect_cell_type("-3.2%") == DataType.PERCENTAGE

    def test_currency_detection(self):
        assert _detect_cell_type("$100") == DataType.CURRENCY
        assert _detect_cell_type("$1,234.56") == DataType.CURRENCY
        assert _detect_cell_type("\u20ac50") == DataType.CURRENCY

    def test_text_detection(self):
        assert _detect_cell_type("hello world") == DataType.TEXT
        assert _detect_cell_type("abc123xyz") == DataType.TEXT


# ---------------------------------------------------------------------------
# Column type resolution tests
# ---------------------------------------------------------------------------


class TestColumnTypeResolution:
    def test_all_same_type(self):
        counts = {DataType.INTEGER.value: 10}
        assert _resolve_column_type(counts) == DataType.INTEGER

    def test_all_empty(self):
        counts = {DataType.EMPTY.value: 5}
        assert _resolve_column_type(counts) == DataType.EMPTY

    def test_empty_dict(self):
        assert _resolve_column_type({}) == DataType.EMPTY

    def test_int_float_becomes_float(self):
        counts = {DataType.INTEGER.value: 5, DataType.FLOAT.value: 3}
        assert _resolve_column_type(counts) == DataType.FLOAT

    def test_date_datetime_becomes_datetime(self):
        counts = {DataType.DATE.value: 2, DataType.DATETIME.value: 8}
        assert _resolve_column_type(counts) == DataType.DATETIME

    def test_dominant_type_wins(self):
        counts = {DataType.TEXT.value: 90, DataType.INTEGER.value: 10}
        assert _resolve_column_type(counts) == DataType.TEXT

    def test_mixed_type(self):
        counts = {
            DataType.TEXT.value: 4,
            DataType.INTEGER.value: 3,
            DataType.EMAIL.value: 3,
        }
        assert _resolve_column_type(counts) == DataType.MIXED

    def test_ignores_empty_in_resolution(self):
        counts = {DataType.INTEGER.value: 5, DataType.EMPTY.value: 100}
        assert _resolve_column_type(counts) == DataType.INTEGER


# ---------------------------------------------------------------------------
# Importer tests
# ---------------------------------------------------------------------------


class TestExcelImporter:
    def test_import_basic(self, sample_xlsx):
        importer = ExcelImporter()
        result = importer.import_file(sample_xlsx)

        assert len(result.sheets) == 1
        assert result.total_rows == 5
        assert result.sheets[0].name == "Employees"
        assert result.sheets[0].col_count == 10
        assert len(result.sheets[0].headers) == 10
        assert result.sheets[0].headers[0] == "ID"

    def test_import_detects_types(self, sample_xlsx):
        importer = ExcelImporter()
        result = importer.import_file(sample_xlsx)
        sheet = result.sheets[0]
        col_types = {c.header: c.detected_type for c in sheet.columns}

        assert col_types["ID"] == DataType.INTEGER
        assert col_types["Name"] == DataType.TEXT
        assert col_types["Email"] == DataType.EMAIL
        assert col_types["Age"] == DataType.INTEGER
        assert col_types["Active"] == DataType.BOOLEAN
        assert col_types["Website"] == DataType.URL
        assert col_types["Phone"] == DataType.PHONE

    def test_import_multi_sheet(self, multi_sheet_xlsx):
        importer = ExcelImporter()
        result = importer.import_file(multi_sheet_xlsx)

        assert len(result.sheets) == 2
        assert result.sheets[0].name == "Products"
        assert result.sheets[1].name == "Financials"
        assert result.total_rows == 7

    def test_import_specific_sheets(self, multi_sheet_xlsx):
        importer = ExcelImporter()
        result = importer.import_file(multi_sheet_xlsx, sheet_names=["Financials"])

        assert len(result.sheets) == 1
        assert result.sheets[0].name == "Financials"
        assert result.total_rows == 3

    def test_import_no_header(self, sample_xlsx):
        importer = ExcelImporter()
        result = importer.import_file(sample_xlsx, has_header=False)

        # With no header, first row becomes data, auto-generated headers
        assert result.sheets[0].row_count == 6
        assert result.sheets[0].headers[0].startswith("Column_")

    def test_import_completeness_tracking(self, sparse_xlsx):
        importer = ExcelImporter()
        result = importer.import_file(sparse_xlsx)
        sheet = result.sheets[0]

        col_a = sheet.columns[0]
        assert col_a.non_empty_count == 3
        assert col_a.empty_count == 2

        col_c = sheet.columns[2]
        assert col_c.non_empty_count == 0
        assert col_c.empty_count == 5
        assert col_c.detected_type == DataType.EMPTY

    def test_import_result_to_dict(self, sample_xlsx):
        importer = ExcelImporter()
        result = importer.import_file(sample_xlsx)
        d = result.to_dict()

        assert d["file_name"].endswith(".xlsx")
        assert d["total_rows"] == 5
        assert d["sheet_count"] == 1
        assert len(d["sheets"]) == 1
        assert len(d["sheets"][0]["columns"]) == 10


# ---------------------------------------------------------------------------
# Analyzer tests
# ---------------------------------------------------------------------------


class TestDataAnalyzer:
    def test_analyze_basic(self, sample_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sample_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)

        assert len(analysis.sheets) == 1
        assert analysis.total_rows == 5
        sa = analysis.sheets[0]
        assert sa.row_count == 5
        assert len(sa.column_stats) == 10

    def test_numeric_stats(self, sample_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sample_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)
        sa = analysis.sheets[0]

        # Find Age column
        age_stats = next(cs for cs in sa.column_stats if cs.header == "Age")
        assert age_stats.min_val == 25
        assert age_stats.max_val == 40
        assert age_stats.mean_val is not None
        assert age_stats.median_val == 30
        assert age_stats.std_dev is not None
        assert age_stats.sum_val == 158  # 30+25+35+28+40

    def test_completeness_calculation(self, sparse_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sparse_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)
        sa = analysis.sheets[0]

        col_a = next(cs for cs in sa.column_stats if cs.header == "A")
        assert col_a.completeness == 60.0  # 3/5

        col_c = next(cs for cs in sa.column_stats if cs.header == "C")
        assert col_c.completeness == 0.0  # 0/5

    def test_duplicate_detection(self, sparse_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sparse_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)
        sa = analysis.sheets[0]

        # Row [1, "hello", None] appears twice
        assert sa.duplicate_rows >= 1

    def test_empty_row_detection(self, sparse_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sparse_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)
        sa = analysis.sheets[0]

        assert sa.completely_empty_rows == 1

    def test_quality_score(self, sample_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sample_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)

        # Full data, no duplicates -> high quality
        assert analysis.overall_quality_score >= 90.0

    def test_low_quality_score(self, sparse_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sparse_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)

        # Sparse data with empties and dupes -> lower quality
        assert analysis.overall_quality_score < 90.0

    def test_top_values(self, sparse_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sparse_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)
        sa = analysis.sheets[0]

        col_b = next(cs for cs in sa.column_stats if cs.header == "B")
        assert len(col_b.top_values) > 0
        # "hello" appears 3 times
        assert col_b.top_values[0][0] == "hello"
        assert col_b.top_values[0][1] == 3

    def test_type_summary(self, sample_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sample_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)
        sa = analysis.sheets[0]

        assert "Integer" in sa.type_summary
        assert "Text" in sa.type_summary
        assert "Email" in sa.type_summary
        assert "Boolean" in sa.type_summary

    def test_boolean_stats(self, sample_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sample_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)
        sa = analysis.sheets[0]

        active_stats = next(cs for cs in sa.column_stats if cs.header == "Active")
        assert active_stats.true_count == 3
        assert active_stats.false_count == 2

    def test_text_stats(self, sample_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sample_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)
        sa = analysis.sheets[0]

        name_stats = next(cs for cs in sa.column_stats if cs.header == "Name")
        assert name_stats.min_length is not None
        assert name_stats.max_length is not None
        assert name_stats.avg_length is not None
        assert name_stats.min_length > 0

    def test_analysis_to_dict(self, sample_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sample_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)
        d = analysis.to_dict()

        assert "overall_quality_score" in d
        assert "sheets" in d
        assert len(d["sheets"]) == 1
        assert "columns" in d["sheets"][0]


# ---------------------------------------------------------------------------
# HTML report tests
# ---------------------------------------------------------------------------


class TestHTMLReportGenerator:
    def test_generate_html_report(self, sample_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sample_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)

        with tempfile.NamedTemporaryFile(suffix=".html", delete=False) as f:
            output_path = f.name

        try:
            gen = HTMLReportGenerator()
            result_path = gen.generate(import_result, analysis, output_path)
            assert result_path == output_path
            assert os.path.isfile(output_path)

            with open(output_path, "r", encoding="utf-8") as f:
                content = f.read()

            assert "Excel Import Report" in content
            assert "Employees" in content
            assert "Integer" in content
            assert "Email" in content
            assert "Boolean" in content
            assert "Data Quality Score" in content
            assert "alice@example.com" in content
        finally:
            os.unlink(output_path)

    def test_html_report_multi_sheet(self, multi_sheet_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(multi_sheet_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)

        with tempfile.NamedTemporaryFile(suffix=".html", delete=False) as f:
            output_path = f.name

        try:
            gen = HTMLReportGenerator()
            gen.generate(import_result, analysis, output_path)

            with open(output_path, "r", encoding="utf-8") as f:
                content = f.read()

            assert "Products" in content
            assert "Financials" in content
        finally:
            os.unlink(output_path)

    def test_html_report_custom_preview(self, sample_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sample_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)

        with tempfile.NamedTemporaryFile(suffix=".html", delete=False) as f:
            output_path = f.name

        try:
            gen = HTMLReportGenerator(preview_rows=2)
            gen.generate(import_result, analysis, output_path)

            with open(output_path, "r", encoding="utf-8") as f:
                content = f.read()

            # Should have limited rows in preview
            assert "first 2 rows" in content
        finally:
            os.unlink(output_path)


# ---------------------------------------------------------------------------
# Excel report tests
# ---------------------------------------------------------------------------


class TestExcelReportGenerator:
    def test_generate_excel_report(self, sample_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sample_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            gen = ExcelReportGenerator()
            result_path = gen.generate(import_result, analysis, output_path)
            assert result_path == output_path
            assert os.path.isfile(output_path)

            # Verify the report is a valid Excel file
            from openpyxl import load_workbook

            wb = load_workbook(output_path, read_only=True)
            sheet_names = wb.sheetnames

            assert "Summary" in sheet_names
            assert any("Analysis" in s for s in sheet_names)
            assert any("Data" in s for s in sheet_names)
            assert any("Quality" in s for s in sheet_names)
            wb.close()
        finally:
            os.unlink(output_path)

    def test_excel_report_summary_content(self, sample_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(sample_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            gen = ExcelReportGenerator()
            gen.generate(import_result, analysis, output_path)

            from openpyxl import load_workbook

            wb = load_workbook(output_path, read_only=True)
            summary = wb["Summary"]

            # Check title
            rows = list(summary.iter_rows(values_only=True))
            title = rows[0][0]
            assert "Excel Import Report" in str(title)
            wb.close()
        finally:
            os.unlink(output_path)

    def test_excel_report_multi_sheet(self, multi_sheet_xlsx):
        importer = ExcelImporter()
        import_result = importer.import_file(multi_sheet_xlsx)
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            gen = ExcelReportGenerator()
            gen.generate(import_result, analysis, output_path)

            from openpyxl import load_workbook

            wb = load_workbook(output_path, read_only=True)
            sheet_names = wb.sheetnames

            # Should have summary + 3 sheets per source sheet (analysis, data, quality)
            assert len(sheet_names) >= 7  # 1 summary + 2*3
            wb.close()
        finally:
            os.unlink(output_path)


# ---------------------------------------------------------------------------
# Integration tests
# ---------------------------------------------------------------------------


class TestIntegration:
    def test_full_pipeline(self, sample_xlsx):
        """Test the complete import -> analyze -> report pipeline."""
        # Import
        importer = ExcelImporter()
        import_result = importer.import_file(sample_xlsx)
        assert len(import_result.sheets) == 1

        # Analyze
        analyzer = DataAnalyzer()
        analysis = analyzer.analyze(import_result)
        assert analysis.overall_quality_score > 0

        # Generate both reports
        with tempfile.TemporaryDirectory() as tmpdir:
            html_path = os.path.join(tmpdir, "report.html")
            excel_path = os.path.join(tmpdir, "report.xlsx")

            html_gen = HTMLReportGenerator()
            html_gen.generate(import_result, analysis, html_path)
            assert os.path.isfile(html_path)
            assert os.path.getsize(html_path) > 1000

            excel_gen = ExcelReportGenerator()
            excel_gen.generate(import_result, analysis, excel_path)
            assert os.path.isfile(excel_path)
            assert os.path.getsize(excel_path) > 1000

    def test_edge_case_single_row(self):
        """Test with a single data row."""
        headers = ["X", "Y"]
        rows = [[1, "hello"]]
        path = _create_test_workbook({"Single": (headers, rows)})

        try:
            importer = ExcelImporter()
            result = importer.import_file(path)
            assert result.total_rows == 1

            analyzer = DataAnalyzer()
            analysis = analyzer.analyze(result)
            assert analysis.sheets[0].column_stats[0].completeness == 100.0

            with tempfile.TemporaryDirectory() as tmpdir:
                html_path = os.path.join(tmpdir, "single.html")
                HTMLReportGenerator().generate(result, analysis, html_path)
                assert os.path.isfile(html_path)
        finally:
            os.unlink(path)

    def test_edge_case_empty_sheet(self):
        """Test with an empty sheet."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        path = tempfile.mktemp(suffix=".xlsx")
        wb.save(path)
        wb.close()

        try:
            importer = ExcelImporter()
            result = importer.import_file(path)
            assert result.total_rows == 0

            analyzer = DataAnalyzer()
            analysis = analyzer.analyze(result)
            assert analysis.overall_quality_score == 0
        finally:
            os.unlink(path)

    def test_large_column_variety(self):
        """Test detection across many data types in one workbook."""
        headers = [
            "Integers",
            "Floats",
            "Booleans",
            "Emails",
            "URLs",
            "Phones",
            "Percentages",
            "Currencies",
            "Text",
            "Dates",
        ]
        rows = [
            [
                1,
                1.5,
                True,
                "a@b.com",
                "https://x.com",
                "+1-555-0000",
                "50%",
                "$100",
                "foo",
                datetime(2024, 1, 1),
            ],
            [
                2,
                2.5,
                False,
                "c@d.com",
                "https://y.com",
                "+1-555-1111",
                "75%",
                "$200",
                "bar",
                datetime(2024, 6, 15),
            ],
            [
                3,
                3.5,
                True,
                "e@f.com",
                "https://z.com",
                "+1-555-2222",
                "90%",
                "$300",
                "baz",
                datetime(2024, 12, 31),
            ],
        ]
        path = _create_test_workbook({"AllTypes": (headers, rows)})

        try:
            importer = ExcelImporter()
            result = importer.import_file(path)
            sheet = result.sheets[0]
            types = {c.header: c.detected_type for c in sheet.columns}

            assert types["Integers"] == DataType.INTEGER
            assert types["Floats"] == DataType.FLOAT
            assert types["Booleans"] == DataType.BOOLEAN
            assert types["Emails"] == DataType.EMAIL
            assert types["URLs"] == DataType.URL
            assert types["Phones"] == DataType.PHONE
            assert types["Percentages"] == DataType.PERCENTAGE
            assert types["Currencies"] == DataType.CURRENCY
            assert types["Text"] == DataType.TEXT
            assert types["Dates"] == DataType.DATETIME

            analyzer = DataAnalyzer()
            analysis = analyzer.analyze(result)
            assert analysis.overall_quality_score == 100.0
            assert len(analysis.sheets[0].type_summary) >= 8

        finally:
            os.unlink(path)
